from openpyxl import load_workbook

wb = load_workbook('functree1.xlsx')
sheet = wb.active
# a1 = sheet['a1'].value
# print(a1)
# print(sheet.cell(4,3).value)

i = 2
j = 1


def findrow(all_row, all_col):
    for row in range(1, all_row + 1):
        for col in range(1, all_col + 1):
            cell = sheet.cell(row, col)
            # print(cell)
            if cell.value is not None:
                print(cell.value)


findrow(28, 6)
