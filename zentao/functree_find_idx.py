from openpyxl import load_workbook

wb = load_workbook('functree2.xlsx')
sheet = wb.active

def get_idxs(start,end):
    indexs =[]
    for i in range(start, end + 1):
        if sheet.cell(i,1).value is not None:
            indexs.append(i)

    indexs.append(end+1)
    print(indexs)
    return indexs


if __name__ == '__main__':
    get_idxs(1,28)



