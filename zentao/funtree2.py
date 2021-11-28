from openpyxl import load_workbook

wb = load_workbook('functree2.xlsx')
sheet = wb.active
# a1 = sheet['a1'].value
# print(a1)
# print(sheet.cell(4,3).value)

i = 2
j = 1

# to_copy = ""


def findrow(all_row, all_col):
    for col in range(1, all_col + 1):
        # for col in range(1,6):
        to_copy = None
        cell_left = None

        for row in range(1, all_row + 1):
            # for row in range(9, 16):
            cell = sheet.cell(row, col)
            print(cell.value)
            # print(cell)

            # 判断有没有 left
            # if int(cell.col_idx) >= 2:
            #     # cell.col_idx
            #     cell_left = sheet.cell(row, col - 1).value

            if cell.value is not None:

                # to_copy = cell.value
                # print(to_copy,"if")
                if sheet.cell(row + 1, col + 1).value is not None :
                    print("当前cell为", sheet.cell(row, col).value)
                    print("右下角为: ", sheet.cell(row + 1, col + 1).value)

                    to_copy = cell.value
                else:
                    to_copy = None


            else:

                # if sheet.cell(row + 1,col).value is None:
                    # cell.value = None
                    cell.value = to_copy
                    # to_copy = None

                # else:
                #     cell.value = None



                    # cell.value = to_copy
                    # print("current cell value:", cell.value)


findrow(28, 6)
wb.save('functree3.xlsx')
wb.close()
