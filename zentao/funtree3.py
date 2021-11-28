from openpyxl import load_workbook


class FuncList:
    def __init__(self, input_file, output_file):
        self.wb = load_workbook(input_file)
        self.sheet = self.wb.active
        self.output_file = output_file
        self.max_orw = self.sheet.max_row
        self.max_column = self.sheet.max_column

    def __del__(self):
        self.wb.close()

    def findrow(self, row_tuple, col_tuple):
        # for col in range(1, all_col + 1):
        for col in range(col_tuple[0], col_tuple[1] + 1):
            # for col in range(1,6):
            to_copy = None

            # for row in range(1, all_row + 1):
            for row in range(row_tuple[0], row_tuple[1] + 1):
                # for row in range(9, 16):
                cell = self.sheet.cell(row, col)
                # print(cell.value)

                if cell.value is not None:

                    if self.sheet.cell(row + 1, col + 1).value is not None:
                        # print("当前cell为", self.sheet.cell(row, col).value)
                        # print("右下角为: ", self.sheet.cell(row + 1, col + 1).value)
                        to_copy = cell.value
                    else:
                        to_copy = None

                elif cell.value is None:
                    cell.value = to_copy
                    # print("current cell value:", cell.value)

        self.wb.save(self.output_file)

    # 大模块分解
    # def get_idxs(self, start, end):
    def get_idxs(self):  # idx_list = [1, 9, 16, 29]
        start = 1
        end = self.max_orw
        indexs = []
        for i in range(start, end + 1):
            if self.sheet.cell(i, 1).value is not None:
                indexs.append(i)

        indexs.append(end + 1)
        print(indexs)
        return indexs

    def get_rowtuples(self):
        idx_list = self.get_idxs()
        row_tuples = []
        for i in range(len(idx_list) - 1):
            row_tuples.append((idx_list[i], idx_list[i + 1] - 1))

        print(row_tuples)
        return row_tuples

    def write_tree(self):
        row_tuples = self.get_rowtuples()
        for row_tuple in row_tuples:
            self.findrow(row_tuple, (1, self.max_column))


if __name__ == '__main__':
    functree = FuncList('functree2.xlsx', 'functree4.xlsx')
    functree.write_tree()

    # findrow((16, 28), (1, 6))
    # 获取大功能列的索引
    # idx_list = functree.get_idxs(1, 28)
    # # 获取每组大功能的列索区间
    # row_tuples = functree.get_rowtuples(idx_list)
    # # 每个功能区间写入功能树列表
    # functree.write_tree(row_tuples)

